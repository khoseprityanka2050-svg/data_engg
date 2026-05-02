"""
Generates docs/30_60_90_execution_plan.xlsx
Run: python docs/generate_execution_plan.py
Requires: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_LAUNCH      = "C6EFCE"   # green — launch-critical
C_LAUNCH_HDR  = "375623"
C_POST        = "FFEB9C"   # amber — post-launch
C_POST_HDR    = "9C5700"
C_PHASE_30    = "D9E1F2"   # light blue
C_PHASE_60    = "FCE4D6"   # light orange
C_PHASE_90    = "E2EFDA"   # light green
C_METRIC_BG   = "F2F2F2"

# ---------------------------------------------------------------------------
# Plan data
# ---------------------------------------------------------------------------

PLAN = [
    # ── DAY 1–30 ──────────────────────────────────────────────────────────
    {
        "phase": "Day 1–30",
        "phase_goal": "Launch-critical foundation: data flowing, decisions traceable",
        "phase_color": C_PHASE_30,
        "items": [
            # (area, task, owner, priority, effort, depends_on, notes)
            ("Infrastructure", "Provision RDS PostgreSQL (Multi-AZ, me-south-1)", "Infra", "Launch-Critical", "2d", "—", "Aurora Serverless v2 target; RDS to start"),
            ("Infrastructure", "Create S3 bronze bucket — KMS-SSE, versioning ON", "Infra", "Launch-Critical", "1d", "—", "Bucket policy: DenyUnencryptedObjectUploads"),
            ("Infrastructure", "Set up MWAA (managed Airflow 2.9)", "Infra", "Launch-Critical", "2d", "—", "One worker node sufficient for 10K/day"),
            ("Infrastructure", "Secrets Manager entries: AECB key, fraud API key, webhook secret, DB DSN", "Infra", "Launch-Critical", "1d", "RDS", "No credentials in code or env vars"),
            ("Database", "Apply bronze schema DDL (4 tables)", "DE", "Launch-Critical", "1d", "RDS", "sql/bronze/create_bronze_tables.sql"),
            ("Database", "Apply silver schema DDL (5 tables + pg_trgm extension)", "DE", "Launch-Critical", "1d", "Bronze DDL", "sql/silver/create_silver_tables.sql"),
            ("Database", "Apply gold schema DDL (decision snapshots + DQ scorecard)", "DE", "Launch-Critical", "1d", "Silver DDL", "sql/gold/*.sql + sql/data_quality/dq_rules.sql"),
            ("Ingestion", "Deploy AECB SFTP ingester — test with sample XML files", "DE", "Launch-Critical", "3d", "Bronze DDL", "Coordinate SFTP credentials with AECB; validate Emirates ID parsing"),
            ("Ingestion", "Deploy Fraud API ingester — integrate with application flow", "DE", "Launch-Critical", "2d", "Bronze DDL", "Confirm API key, rate limits, and SLA with provider"),
            ("Ingestion", "Deploy AML webhook receiver behind API Gateway", "DE", "Launch-Critical", "2d", "Bronze DDL", "Set idempotency; validate webhook secret rotation plan"),
            ("Ingestion", "Deploy customer profile CDC extractor — full load first", "DE", "Launch-Critical", "2d", "Bronze DDL", "Confirm updated_at index exists on source PostgreSQL"),
            ("Identity Resolution", "Run silver/identity_resolution.sql — validate against test data", "DE", "Launch-Critical", "2d", "All ingestion", "Test cross-source matching with real Emirates IDs"),
            ("Identity Resolution", "Tune pg_trgm threshold on Arabic names (sample 500 records)", "DE", "Launch-Critical", "2d", "Identity resolution", "Arabic transliterations require lower threshold — document exceptions"),
            ("Decisioning", "Wire gold.create_decision_snapshot() into application API", "DE+Eng", "Launch-Critical", "3d", "Silver tables", "Every credit decision must call this function before returning a result"),
            ("Data Quality", "Activate DQ scorecard engine — MUST_PASS rules only", "DE", "Launch-Critical", "2d", "Gold DDL", "Connect SNS → team Slack for MUST_PASS failures"),
            ("Orchestration", "Deploy Airflow DAG — run end-to-end with real data", "DE", "Launch-Critical", "2d", "All above", "Validate short-circuit gate on DQ failures"),
            ("Compliance", "Verify audit chain: gold FK → silver FK → bronze raw payload", "DE+Legal", "Launch-Critical", "1d", "Gold DDL", "Write a sample audit query; review with compliance officer"),
            ("Testing", "Load test at 10K decisions/day sustained for 4 hours", "DE+QA", "Launch-Critical", "2d", "Full pipeline", "Measure p95 latency on snapshot insert; index if > 200ms"),
        ],
    },
    # ── DAY 31–60 ─────────────────────────────────────────────────────────
    {
        "phase": "Day 31–60",
        "phase_goal": "Stabilise, monitor, and harden for growth to 30K/day",
        "phase_color": C_PHASE_60,
        "items": [
            ("Data Quality", "Activate HIGH and MEDIUM DQ rules", "DE", "Post-Launch", "2d", "Day 30 DQ", "Set alert thresholds based on first 30 days of actual data"),
            ("Data Quality", "Build gold.v_dq_alerts dashboard in CloudWatch or Metabase", "DE", "Post-Launch", "2d", "DQ rules live", "Risk team needs daily DQ health view"),
            ("Portfolio Monitoring", "Activate portfolio_daily_snapshot refresh (daily CALL)", "DE", "Post-Launch", "1d", "Day 30 pipeline", "Schedule as separate Airflow task after gold refresh"),
            ("Portfolio Monitoring", "Build approval rate trend dashboard (v_approval_rate_trend)", "Analytics", "Post-Launch", "3d", "Snapshot live", "Connect Metabase or QuickSight to gold schema read replica"),
            ("Portfolio Monitoring", "Build risk concentration matrix (v_risk_concentration)", "Analytics", "Post-Launch", "2d", "Snapshot live", "Cross-tab: score band × DTI band — CRO weekly review"),
            ("Identity Resolution", "Review identity_resolution_exceptions backlog weekly", "DE", "Post-Launch", "Ongoing", "Day 30 pipeline", "Set up weekly Airflow task to email count + sample"),
            ("Reliability", "Add dead-letter queue for failed AML webhook deliveries", "DE+Infra", "Post-Launch", "2d", "Webhook live", "Provider retries 3×; DLQ catches remainder for manual reprocess"),
            ("Reliability", "Implement Airflow SLA missed alerts (DAG must complete by 06:00)", "DE", "Post-Launch", "1d", "Airflow DAG", "Pipeline must complete before business opens at 09:00 UAE"),
            ("Reliability", "Add RDS CloudWatch alarms: CPU > 70%, free storage < 20 GB", "Infra", "Post-Launch", "1d", "RDS live", "Auto-scaling storage enabled; alarm before it's needed"),
            ("Security", "Column-level encryption for Emirates ID and DOB in silver/gold", "DE+Sec", "Post-Launch", "4d", "Silver DDL", "Use pgcrypto or AWS RDS encryption + application-layer masking"),
            ("Security", "Implement row-level security: analysts can only see their product", "DE+Sec", "Post-Launch", "2d", "Silver DDL", "RLS policies on gold.credit_decision_inputs by product_type"),
            ("Performance", "Add partial indexes on gold for common filter patterns", "DE", "Post-Launch", "1d", "30d of data", "e.g. WHERE decision_outcome = 'APPROVED' AND decided_at > NOW()-30d"),
            ("Performance", "Profile identity resolution queries; add GIN index on full_name if needed", "DE", "Post-Launch", "2d", "30d of data", "pg_trgm GIN index accelerates similarity() lookups 10–50×"),
            ("Documentation", "Write data dictionary for all silver and gold tables", "DE", "Post-Launch", "3d", "Schema stable", "Required for CBUAE audit; include data lineage notes"),
            ("Process", "Weekly DQ review cadence with risk team (30-min meeting)", "DE+Risk", "Post-Launch", "Ongoing", "DQ live", "Review dq_scorecard trends; adjust thresholds as needed"),
        ],
    },
    # ── DAY 61–90 ─────────────────────────────────────────────────────────
    {
        "phase": "Day 61–90",
        "phase_goal": "Scale to 100K/day, analytics layer, and CBUAE readiness",
        "phase_color": C_PHASE_90,
        "items": [
            ("Scale", "Migrate gold analytics queries to S3 Parquet + Athena", "DE+Infra", "Post-Launch", "5d", "Day 60 stable", "30K+/day makes RDS analytics queries expensive; Athena = pay-per-scan"),
            ("Scale", "Automate Parquet export: gold tables → S3 daily via Airflow", "DE", "Post-Launch", "3d", "Athena ready", "Partition by product_type/snapshot_date for efficient pruning"),
            ("Scale", "Upgrade RDS to Aurora Serverless v2 (auto-scales to 100K/day)", "Infra", "Post-Launch", "3d", "Load test", "Blue/green deployment; < 1 min failover; no downtime"),
            ("Scale", "Load test at 100K decisions/day (simulate with Locust)", "DE+QA", "Post-Launch", "3d", "Aurora upgrade", "Acceptance: p95 snapshot insert < 500ms, DQ run < 10 min"),
            ("Analytics", "Build CBUAE regulatory report (auto-generated from gold)", "DE+Compliance", "Post-Launch", "5d", "Gold stable", "Monthly report: approvals by nationality, product, DTI band"),
            ("Analytics", "Set up Redshift Serverless for cross-product analytics (optional)", "Infra", "Post-Launch", "4d", "Parquet on S3", "If Athena query latency > 10s on complex joins; evaluate vs cost"),
            ("Fraud", "Build velocity monitoring: >3 applications/customer/7 days alert", "DE+Fraud", "Post-Launch", "3d", "Gold 30d data", "Fraud rings often submit multiple applications with slight variations"),
            ("Fraud", "Integrate fraud score trend: flag rising scores on existing customers", "DE", "Post-Launch", "2d", "60d fraud data", "Silver fraud_scores.customer_key allows longitudinal view"),
            ("ML Readiness", "Export feature store snapshot from silver for model training", "DE+DS", "Post-Launch", "4d", "Silver stable", "Train initial scorecard model on first 60 days of decisions"),
            ("ML Readiness", "Add model_version column to gold.credit_decision_inputs", "DE", "Post-Launch", "1d", "Gold stable", "Needed to A/B test scorecard models in production"),
            ("Compliance", "Full CBUAE audit trail dry-run: trace 50 random decisions end-to-end", "DE+Legal", "Post-Launch", "2d", "All layers", "Document findings; resolve any gaps before formal audit"),
            ("Compliance", "Implement data access log (who queried what PII, when)", "Infra+DE", "Post-Launch", "3d", "RLS live", "CloudTrail + RDS audit log → CloudWatch; 1-year retention"),
            ("Reliability", "Implement cross-region S3 replication for bronze bucket", "Infra", "Post-Launch", "1d", "S3 stable", "DR requirement: RPO < 24h; RTO < 4h for bronze data"),
            ("Process", "Runbook for common failure modes (stale AECB, webhook gap)", "DE", "Post-Launch", "3d", "60d operations", "On-call engineers need step-by-step remediation guides"),
        ],
    },
]

# ---------------------------------------------------------------------------
# Success metrics per phase
# ---------------------------------------------------------------------------

METRICS = [
    ("Day 30 Success Criteria", [
        "All 4 bronze sources ingesting daily without manual intervention",
        "100% of credit decisions have an entry in gold.credit_decision_inputs",
        "DQ MUST_PASS rules: 0 failures in 7 consecutive days",
        "Audit chain query returns complete lineage for 100% of approved decisions",
        "Pipeline completes by 06:00 UAE time for 10K decisions/day",
    ]),
    ("Day 60 Success Criteria", [
        "DQ HIGH rules: pass rate ≥ threshold for 95% of daily runs",
        "Portfolio dashboard live and reviewed by risk team weekly",
        "Identity resolution exception rate < 2% of daily applications",
        "Zero PII data accessible without RLS enforcement",
        "Incident response: DQ alert to resolution < 2 hours (P1)",
    ]),
    ("Day 90 Success Criteria", [
        "Pipeline handles 100K decisions/day: p95 snapshot insert < 500ms",
        "CBUAE audit dry-run: 100% of sampled decisions fully traceable",
        "Athena analytics query latency < 10s on 90-day lookback",
        "Monthly CBUAE regulatory report auto-generated with 0 manual edits",
        "Feature store available for model training on full 90-day dataset",
    ]),
]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def make_border(style="thin") -> Border:
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row: int, values: list, bg: str, fg: str = "FFFFFF", size: int = 10):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = make_fill(bg)
        cell.font = make_font(bold=True, color=fg, size=size)
        cell.border = make_border()
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def write_plan(output_path: str):
    wb = Workbook()

    # ── Sheet 1: Execution Plan ──────────────────────────────────────────
    ws = wb.active
    ws.title = "30-60-90 Execution Plan"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Mal Credit Decision Platform — 30/60/90 Day Execution Plan"
    title_cell.font = Font(bold=True, size=16, color=C_HEADER_FG, name="Calibri")
    title_cell.fill = make_fill(C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Legend
    ws.merge_cells("A2:H2")
    legend = ws["A2"]
    legend.value = "  Launch-Critical = must be live on Day 1 of product launch    |    Post-Launch = important but does not block revenue"
    legend.font = Font(size=9, italic=True, color="444444", name="Calibri")
    legend.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column widths
    col_widths = [18, 52, 12, 16, 8, 20, 38]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Area", "Task", "Owner", "Priority", "Effort", "Depends On", "Notes"]
    current_row = 3

    for phase in PLAN:
        # Phase heading
        ws.merge_cells(f"A{current_row}:G{current_row}")
        phase_cell = ws.cell(row=current_row, column=1)
        phase_cell.value = f"  {phase['phase']} — {phase['phase_goal']}"
        phase_cell.fill = make_fill(C_HEADER_BG)
        phase_cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        phase_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column headers
        apply_header(ws, current_row, headers, bg="2F5496", size=9)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for item in phase["items"]:
            area, task, owner, priority, effort, depends, notes = item
            bg = C_LAUNCH if priority == "Launch-Critical" else C_POST
            row_vals = [area, task, owner, priority, effort, depends, notes]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.fill = make_fill(bg)
                cell.border = make_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9, name="Calibri",
                                 bold=(col == 4 and priority == "Launch-Critical"))
            ws.row_dimensions[current_row].height = 30
            current_row += 1

        current_row += 1  # spacer

    # ── Sheet 2: Success Metrics ─────────────────────────────────────────
    ws2 = wb.create_sheet("Success Metrics")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 78

    ws2.merge_cells("A1:B1")
    m_title = ws2["A1"]
    m_title.value = "Success Criteria by Phase"
    m_title.font = Font(bold=True, size=14, color=C_HEADER_FG, name="Calibri")
    m_title.fill = make_fill(C_HEADER_BG)
    m_title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    mrow = 2
    phase_bgs = [C_PHASE_30, C_PHASE_60, C_PHASE_90]
    for (phase_label, criteria), bg in zip(METRICS, phase_bgs):
        ws2.merge_cells(f"A{mrow}:B{mrow}")
        h = ws2.cell(row=mrow, column=1, value=phase_label)
        h.fill = make_fill(C_HEADER_BG)
        h.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        h.alignment = Alignment(vertical="center")
        ws2.row_dimensions[mrow].height = 22
        mrow += 1

        for criterion in criteria:
            ws2.cell(row=mrow, column=1, value="✓ Metric").font = Font(bold=True, size=9, name="Calibri")
            ws2.cell(row=mrow, column=1).fill = make_fill(bg)
            ws2.cell(row=mrow, column=1).border = make_border()
            ws2.cell(row=mrow, column=1).alignment = Alignment(horizontal="center", vertical="center")

            c = ws2.cell(row=mrow, column=2, value=criterion)
            c.fill = make_fill(C_METRIC_BG)
            c.border = make_border()
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.font = Font(size=9, name="Calibri")
            ws2.row_dimensions[mrow].height = 22
            mrow += 1

        mrow += 1

    # ── Sheet 3: RACI ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("RACI")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 40
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws3.column_dimensions[col_letter].width = 14

    ws3.merge_cells("A1:F1")
    r_title = ws3["A1"]
    r_title.value = "RACI — Responsibility Matrix"
    r_title.font = Font(bold=True, size=14, color=C_HEADER_FG, name="Calibri")
    r_title.fill = make_fill(C_HEADER_BG)
    r_title.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    raci_headers = ["Workstream", "Data Engineering", "Infra/DevOps", "Analytics", "Risk/Compliance", "Engineering"]
    apply_header(ws3, 2, raci_headers, bg="2F5496", size=9)
    ws3.row_dimensions[2].height = 20

    raci_data = [
        ("Bronze ingestion (all 4 sources)",    "R/A", "C",   "I",   "I",   "C"),
        ("Identity resolution logic",            "R/A", "I",   "C",   "C",   "I"),
        ("Gold decision snapshot function",      "R/A", "I",   "I",   "I",   "C"),
        ("Data quality rules & thresholds",      "R",   "I",   "I",   "A",   "I"),
        ("Portfolio monitoring mart",            "R",   "I",   "A",   "C",   "I"),
        ("AWS infrastructure provisioning",      "C",   "R/A", "I",   "I",   "I"),
        ("Airflow DAG orchestration",            "R/A", "C",   "I",   "I",   "I"),
        ("CBUAE compliance audit trail",         "C",   "I",   "I",   "R/A", "I"),
        ("Fraud API & AML provider integration", "R",   "C",   "I",   "A",   "C"),
        ("Analytics dashboards",                 "C",   "I",   "R/A", "C",   "I"),
        ("Security & PII encryption",            "C",   "R",   "I",   "A",   "C"),
        ("Load testing & performance",           "R",   "C",   "I",   "I",   "A"),
    ]
    note_row = ["R = Responsible · A = Accountable · C = Consulted · I = Informed", "", "", "", "", ""]
    legend_fill = make_fill("F2F2F2")

    for rrow, (workstream, *roles) in enumerate(raci_data, start=3):
        bg = "FFFFFF" if rrow % 2 == 0 else "F7F9FC"
        ws3.cell(row=rrow, column=1, value=workstream).fill = make_fill(bg)
        ws3.cell(row=rrow, column=1).font = Font(size=9, name="Calibri")
        ws3.cell(row=rrow, column=1).border = make_border()
        for col_i, role in enumerate(roles, start=2):
            cell = ws3.cell(row=rrow, column=col_i, value=role)
            cell.fill = make_fill("C6EFCE" if "A" in role else bg)
            cell.font = Font(size=9, bold=("A" in role), name="Calibri")
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[rrow].height = 18

    legend_row = len(raci_data) + 4
    ws3.merge_cells(f"A{legend_row}:F{legend_row}")
    l = ws3.cell(row=legend_row, column=1, value=note_row[0])
    l.font = Font(italic=True, size=8, name="Calibri", color="666666")
    l.alignment = Alignment(horizontal="center")

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "30_60_90_execution_plan.xlsx")
    write_plan(out)
